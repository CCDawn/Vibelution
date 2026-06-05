from pathlib import Path

from openpyxl import load_workbook


path = Path(__file__).resolve().parent / "outputs" / "挑战杯报名信息与赛题提取表.xlsx"
workbook = load_workbook(path)
worksheet = workbook["报名填报信息"]

updates = {
    "揭榜作品名称": "基于国产开源大模型的脑语言机制启发神经网络算法发现平台",
    "初期研究情况": (
        "已围绕人类语言中枢机制与人工神经网络算法之间的关系开展初步调研，"
        "重点关注语言理解、语义加工、句法组合、语境推理等脑语言功能与Transformer、"
        "注意力机制、表征学习等神经网络方法之间的关联。"
        "项目拟基于Qwen系列国产开源大模型和多智能体系统，构建文献挖掘、"
        "机制建模、算法假设生成、实验验证与迭代优化流程，"
        "探索受人类语言中枢启发的有效神经网络算法。"
    ),
}


def normalize(value):
    return str(value or "").replace("\r\n", "\n").replace("\n", "").replace(" ", "").strip()


changed = []
for row in range(1, min(worksheet.max_row, 10) + 1):
    for col in range(1, worksheet.max_column + 1):
        key = normalize(worksheet.cell(row, col).value)
        for header, value in updates.items():
            if key == normalize(header):
                target = worksheet.cell(row + 1, col)
                target.value = value
                changed.append((header, target.coordinate))

if len(changed) != len(updates):
    raise RuntimeError(f"Expected {len(updates)} updates, got {changed}")

workbook.save(path)
print(changed)
